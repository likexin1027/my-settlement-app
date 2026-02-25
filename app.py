import io
import pandas as pd
import streamlit as st
import requests  

st.set_page_config(page_title="101ä¿±ä¹éƒ¨ç»“ç®—å·¥å…·", layout="wide")

def normalize_platform(s):
    if pd.isna(s):
        return ""
    x = str(s).strip().lower()
    if "bç«™" in x or "bilibili" in x or "å“”å“©" in x:
        return "Bç«™"
    if "å°çº¢ä¹¦" in x or "red" in x:
        return "å°çº¢ä¹¦"
    if "è§†é¢‘å·" in x:
        return "è§†é¢‘å·"
    if "æŠ–éŸ³" in x or "douyin" in x:
        return "æŠ–éŸ³"
    return s

def parse_number(v):
    if pd.isna(v):
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s.endswith("+"):
        s = s[:-1]
    m = 1.0
    if "äº¿" in s:
        s = s.replace("äº¿", "")
        m = 100000000.0
    elif "w" in s.lower():
        s = s.lower().replace("w", "")
        m = 10000.0
    elif "ä¸‡" in s:
        s = s.replace("ä¸‡", "")
        m = 10000.0
    try:
        return float(s) * m
    except:
        try:
            return float(s)
        except:
            return 0.0

def create_default_mapping():
    thresholds = [1000000, 500000, 200000, 100000, 50000, 30000, 10000]
    labels = ["â‰¥100w", "â‰¥50w", "â‰¥20w", "â‰¥10w", "â‰¥5w", "â‰¥3w", "â‰¥1w"]
    rows = []
    for t, lab in zip(thresholds, labels):
        for plat in ["Bç«™", "å°çº¢ä¹¦", "æŠ–éŸ³", "è§†é¢‘å·"]:
            rows.append({"å¹³å°": plat, "é˜ˆå€¼æ ‡ç­¾": lab, "é˜ˆå€¼æ•°å€¼": t, "å¥–åŠ±é‡‘é¢": 0.0})
    df = pd.DataFrame(rows)
    return df

def load_default_mapping():
    try:
        try:
            mdf = pd.read_csv("å¥–åŠ±é‡‘é¢t.csv")
        except:
            mdf = pd.read_csv("å¥–åŠ±é‡‘é¢t.csv", encoding="gbk")
        cols = set(mdf.columns)
        need_plat = "å¹³å°" in cols
        has_val = "é˜ˆå€¼æ•°å€¼" in cols
        has_lab = "é˜ˆå€¼æ ‡ç­¾" in cols
        has_amt = "å¥–åŠ±é‡‘é¢" in cols
        if need_plat and has_amt and (has_val or has_lab):
            out = mdf.copy()
            if not has_val and has_lab:
                out["é˜ˆå€¼æ•°å€¼"] = out["é˜ˆå€¼æ ‡ç­¾"].apply(normalize_label_to_value)
            if not has_lab and has_val:
                out["é˜ˆå€¼æ ‡ç­¾"] = out["é˜ˆå€¼æ•°å€¼"].apply(value_to_label)
            out = out[out["å¹³å°"].isin(["Bç«™", "å°çº¢ä¹¦", "æŠ–éŸ³", "è§†é¢‘å·"])]
            out = out[["å¹³å°", "é˜ˆå€¼æ ‡ç­¾", "é˜ˆå€¼æ•°å€¼", "å¥–åŠ±é‡‘é¢"]].dropna(subset=["é˜ˆå€¼æ•°å€¼", "å¥–åŠ±é‡‘é¢"])
            return out
        return create_default_mapping()
    except:
        return create_default_mapping()

def normalize_label_to_value(lab):
    if pd.isna(lab):
        return None
    s = str(lab).strip().lower().replace("â‰¥", "").replace("+", "")
    s = s.replace("ä¸‡", "w")
    if "w" in s:
        try:
            n = float(s.replace("w", ""))
            return int(n * 10000)
        except:
            return None
    try:
        return int(float(s))
    except:
        return None

def value_to_label(v):
    if v >= 1000000:
        return "â‰¥100w"
    if v >= 500000:
        return "â‰¥50w"
    if v >= 200000:
        return "â‰¥20w"
    if v >= 100000:
        return "â‰¥10w"
    if v >= 50000:
        return "â‰¥5w"
    if v >= 30000:
        return "â‰¥3w"
    if v >= 10000:
        return "â‰¥1w"
    return f"â‰¥{int(v)}"

def build_reward_lookup(df):
    d = {}
    for plat in df["å¹³å°"].unique():
        sub = df[df["å¹³å°"] == plat].sort_values("é˜ˆå€¼æ•°å€¼", ascending=False)
        d[plat] = [(row["é˜ˆå€¼æ•°å€¼"], float(row["å¥–åŠ±é‡‘é¢"])) for _, row in sub.iterrows()]
    return d

def describe_excel_error(err, filename):
    s = str(err).lower()
    reasons = []
    if "encrypted" in s or "password" in s:
        reasons.append("æ–‡ä»¶åŠ å¯†æˆ–å—ä¿æŠ¤")
    if "not a zip file" in s or "unsupported file format" in s or "badzipfile" in s:
        reasons.append("æ–‡ä»¶æŸåæˆ–å¹¶éæ ‡å‡†xlsx/xls")
    if "calamine" in s and ("not installed" in s or "module" in s):
        reasons.append("ç¼ºå°‘è¯»å–å¼•æ“ï¼Œè¯·å®‰è£…python-calamine")
    if "openpyxl" in s and ("styles" in s or "fills" in s):
        reasons.append("å¤æ‚æ ·å¼å¯¼è‡´è§£æå¤±è´¥ï¼Œå»ºè®®é‡å¯¼å‡ºæˆ–ç®€åŒ–æ ·å¼")
    if filename.endswith(".xls") and ("xlrd" in s or "format" in s):
        reasons.append(".xlså…¼å®¹æ€§é—®é¢˜ï¼Œå»ºè®®å¦å­˜ä¸º.xlsxåå†ä¸Šä¼ ")
    if "filetype" in s or "content-type" in s:
        reasons.append("æ‰©å±•åä¸å®é™…å†…å®¹ä¸åŒ¹é…")
    msg = "Excelè¯»å–å¤±è´¥"
    if reasons:
        msg += "ï¼š" + "ï¼›".join(reasons)
    msg += f"ã€‚åŸå§‹ä¿¡æ¯ï¼š{str(err)}"
    return msg

def read_xlsx_robust(bio):
    try:
        return pd.read_excel(bio, engine="calamine")
    except:
        bio.seek(0)
        try:
            return pd.read_excel(bio, engine="openpyxl")
        except:
            bio.seek(0)
            try:
                from openpyxl import load_workbook
                wb = load_workbook(bio, data_only=True, read_only=True)
                ws = wb.active
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
                if not data:
                    return pd.DataFrame()
                header = [str(x) if x is not None else "" for x in data[0]]
                rows = data[1:]
                return pd.DataFrame(rows, columns=header)
            except Exception as e:
                raise e

def base_reward(plat, views, lookup):
    if plat not in lookup:
        return 0.0
    for th, amt in lookup[plat]:
        if views >= th:
            return amt
    return 0.0

def limited_time_bonus(views, typ):
    if views > 10000 and isinstance(typ, str):
        s = typ.lower()
        if ("çƒ­ç‚¹æ¨è" in s) or ("æ–°æ˜¥ä¸»é¢˜" in s):
            return 50.0
    return 0.0

def excellence_bonus(plat, typ, likes, views):
    b = 0.0
    if isinstance(typ, str):
        s = typ
        if plat == "Bç«™":
            if "çƒ­æœ" in s:
                b += 100.0
            if "çƒ­é—¨" in s:
                b += 200.0
        if "çŸ­è§†é¢‘" in s and likes >= 100000:
            b += 300.0
        if "çŸ­è§†é¢‘" in s and views >= 2000000:
            b += 1000.0
    return b

def pick_top5_per_author(df):
    df = df.copy()
    df["æ˜¯å¦è®¡å…¥ç»“ç®—"] = False
    pos = df["æ€»å¥–åŠ±"] > 0
    for author, group in df[pos].groupby("è´¦å·åç§°"):
        idx = group.sort_values("æ€»å¥–åŠ±", ascending=False).head(5).index
        df.loc[idx, "æ˜¯å¦è®¡å…¥ç»“ç®—"] = True
    return df

def filter_banned(df, text_cols):
    banned = ["BUG", "å»ºè®®", "æ‹‰è¸©"]
    mask = pd.Series([False] * len(df))
    for col in text_cols:
        if col in df.columns:
            s = df[col].astype(str)
            for w in banned:
                mask = mask | s.str.contains(w, case=False, na=False)
    out = df.copy()
    out["æ’é™¤åŸå› "] = ""
    out.loc[mask, "æ’é™¤åŸå› "] = "åŒ…å«æ•æ„Ÿè¯"
    return out[~mask], out[mask]

def render():
    st.title("101ä¿±ä¹éƒ¨æ´»åŠ¨å¥–é‡‘ç»“ç®—")
    st.caption("ä¸Šä¼ æ•°æ®ï¼Œé…ç½®åŸºç¡€å¥–åŠ±ï¼Œè‡ªåŠ¨è®¡ç®—é™æ—¶ä¸ä¼˜ç§€å¥–åŠ±ï¼ŒæŒ‰ä½œè€…é™é¢è¾“å‡ºç»“ç®—ç»“æœ")
    tabs = st.tabs(["ç»“ç®—ä¸­å¿ƒ", "è§„åˆ™è®¾ç½®"])
    with tabs[1]:
        mapping = load_default_mapping()
        cfg = st.file_uploader("ä¸Šä¼ å¥–åŠ±é…ç½®ï¼ˆExcel/CSVï¼‰", type=["xlsx", "xls", "csv"], key="cfg")
        if cfg is not None:
            n = getattr(cfg, "name", "").lower()
            try:
                if n.endswith(".csv"):
                    try:
                        mdf = pd.read_csv(cfg)
                    except:
                        cfg.seek(0)
                        mdf = pd.read_csv(cfg, encoding="gbk")
                else:
                    data = cfg.read()
                    bio = io.BytesIO(data)
                    if n.endswith(".xlsx"):
                        try:
                            mdf = pd.read_excel(bio, engine="calamine")
                        except:
                            bio.seek(0)
                            mdf = pd.read_excel(bio, engine="openpyxl")
                    elif n.endswith(".xls"):
                        mdf = pd.read_excel(bio, engine="xlrd")
                    else:
                        bio.seek(0)
                        mdf = pd.read_excel(bio)
            except:
                mdf = None
            if mdf is not None:
                cols = set(mdf.columns)
                need_plat = "å¹³å°" in cols
                has_val = "é˜ˆå€¼æ•°å€¼" in cols
                has_lab = "é˜ˆå€¼æ ‡ç­¾" in cols
                has_amt = "å¥–åŠ±é‡‘é¢" in cols
                if need_plat and has_amt and (has_val or has_lab):
                    out = mdf.copy()
                    if not has_val and has_lab:
                        out["é˜ˆå€¼æ•°å€¼"] = out["é˜ˆå€¼æ ‡ç­¾"].apply(normalize_label_to_value)
                    if not has_lab and has_val:
                        out["é˜ˆå€¼æ ‡ç­¾"] = out["é˜ˆå€¼æ•°å€¼"].apply(value_to_label)
                    out = out[out["å¹³å°"].isin(["Bç«™", "å°çº¢ä¹¦", "æŠ–éŸ³", "è§†é¢‘å·"])]
                    out = out[["å¹³å°", "é˜ˆå€¼æ ‡ç­¾", "é˜ˆå€¼æ•°å€¼", "å¥–åŠ±é‡‘é¢"]].dropna(subset=["é˜ˆå€¼æ•°å€¼", "å¥–åŠ±é‡‘é¢"])
                    mapping = out
        mapping = st.data_editor(mapping, num_rows="dynamic", width="stretch")
    with tabs[0]:
        uploaded = st.file_uploader("ä¸Šä¼ Excelæˆ–CSVæ–‡ä»¶", type=["xlsx", "xls", "csv"])
    lookup = build_reward_lookup(mapping)
    if uploaded is None:
        return
    name = getattr(uploaded, "name", "").lower()
    if name.endswith(".csv"):
        try:
            df = pd.read_csv(uploaded)
        except:
            try:
                uploaded.seek(0)
                df = pd.read_csv(uploaded, encoding="gbk")
            except:
                st.error("CSVè¯»å–å¤±è´¥ï¼Œè¯·ç¡®è®¤ç¼–ç ä¸æ–‡ä»¶æ ¼å¼")
                return
    else:
        try:
            data = uploaded.read()
            bio = io.BytesIO(data)
            if name.endswith(".xlsx"):
                df = read_xlsx_robust(bio)
            elif name.endswith(".xls"):
                df = pd.read_excel(bio, engine="xlrd")
            else:
                bio.seek(0)
                df = pd.read_excel(bio)
        except Exception as e:
            st.error(describe_excel_error(e, name))
            return
    required = ["æ¸ é“", "æ’­æ”¾é‡", "ç‚¹èµ", "ä½œå“ç±»å‹", "è´¦å·åç§°"]
    miss = [c for c in required if c not in df.columns]
    if miss:
        st.error("ç¼ºå°‘å­—æ®µ: " + ", ".join(miss))
        return
    df["æ¸ é“"] = df["æ¸ é“"].apply(normalize_platform)
    df["æ’­æ”¾é‡æ•°å€¼"] = df["æ’­æ”¾é‡"].apply(parse_number)
    df["ç‚¹èµæ•°å€¼"] = df["ç‚¹èµ"].apply(parse_number)
    text_cols = []
    for c in ["ä½œå“ç±»å‹", "å†…å®¹", "æ ‡é¢˜", "ä½œå“æ ‡é¢˜"]:
        if c in df.columns:
            text_cols.append(c)
    kept, removed = filter_banned(df, text_cols if text_cols else ["ä½œå“ç±»å‹"])
    kept["åŸºç¡€å¥–åŠ±"] = kept.apply(lambda x: base_reward(x["æ¸ é“"], x["æ’­æ”¾é‡æ•°å€¼"], lookup), axis=1)
    kept["é™æ—¶å¥–åŠ±"] = kept.apply(lambda x: limited_time_bonus(x["æ’­æ”¾é‡æ•°å€¼"], x["ä½œå“ç±»å‹"]), axis=1)
    kept["ä¼˜ç§€å¥–åŠ±"] = kept.apply(lambda x: excellence_bonus(x["æ¸ é“"], x["ä½œå“ç±»å‹"], x["ç‚¹èµæ•°å€¼"], x["æ’­æ”¾é‡æ•°å€¼"]), axis=1)
    kept["æ€»å¥–åŠ±"] = kept[["åŸºç¡€å¥–åŠ±", "é™æ—¶å¥–åŠ±", "ä¼˜ç§€å¥–åŠ±"]].sum(axis=1)
    kept = pick_top5_per_author(kept)
    result = kept.copy()
    result = result[["æ¸ é“", "è´¦å·åç§°", "æ’­æ”¾é‡", "ç‚¹èµ", "ä½œå“ç±»å‹", "åŸºç¡€å¥–åŠ±", "é™æ—¶å¥–åŠ±", "ä¼˜ç§€å¥–åŠ±", "æ€»å¥–åŠ±", "æ˜¯å¦è®¡å…¥ç»“ç®—"]]
    with tabs[0]:
        summary = result[result["æ˜¯å¦è®¡å…¥ç»“ç®—"]].groupby("è´¦å·åç§°", as_index=False).agg({ "æ€»å¥–åŠ±": "sum","æ’­æ”¾é‡æ•°å€¼": "sum"}).rename(columns={"æ€»å¥–åŠ±": "ç»“ç®—é‡‘é¢", "æ’­æ”¾é‡æ•°å€¼": "æ€»æ’­æ”¾é‡"})

# å­˜å…¥ç¼“å­˜ï¼Œç»™ AI çœ‹
        st.session_state["summary_data"] = summary
        total_payout = summary["ç»“ç®—é‡‘é¢"].sum() if not summary.empty else 0.0
        total_views = result[result["æ˜¯å¦è®¡å…¥ç»“ç®—"]]["æ’­æ”¾é‡æ•°å€¼"].sum() if "æ’­æ”¾é‡æ•°å€¼" in result.columns else 0.0
        counted = int(result["æ˜¯å¦è®¡å…¥ç»“ç®—"].sum())
        authors = summary.shape[0]
        cols = st.columns(4)
        cols[0].metric("æ€»ç»“ç®—é‡‘é¢", f"{total_payout:,.2f} å…ƒ")
        cols[1].metric("æ€»æ’­æ”¾é‡", f"{int(total_views):,}")
        cols[2].metric("è®¡å…¥æ¡ç›®æ•°", f"{counted}")
        cols[3].metric("ä½œè€…æ•°", f"{authors}")
        st.subheader("ç»“ç®—é¢„è§ˆ")
        st.dataframe(result, width="stretch")
        st.subheader("å¥–é‡‘Top5ä½œè€…")
        top5 = summary.sort_values("ç»“ç®—é‡‘é¢", ascending=False).head(5)
        st.bar_chart(top5.set_index("è´¦å·åç§°"))
        st.subheader("è¢«æ’é™¤å†…å®¹")
        if not removed.empty:
            st.dataframe(removed, width="stretch")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="ç»“ç®—æ˜ç»†")
        summary.to_excel(writer, index=False, sheet_name="ä½œè€…æ±‡æ€»")
        mapping.to_excel(writer, index=False, sheet_name="å¥–åŠ±é…ç½®")
    st.download_button("ä¸‹è½½å¤„ç†åçš„Excel", data=buffer.getvalue(), file_name="101ä¿±ä¹éƒ¨ç»“ç®—ç»“æœ.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.divider()
    st.subheader(" 101 ç»“ç®—æ™ºèƒ½åŠ©æ‰‹")

    # æ£€æŸ¥æ˜¯å¦æœ‰è®¡ç®—å¥½çš„æ•°æ®
    if "summary_data" in st.session_state and st.session_state["summary_data"] is not None:
        summary_for_ai = st.session_state["summary_data"]
        context_text = summary_for_ai.to_string(index=False)
        
        # åˆå§‹åŒ–æ¶ˆæ¯è®°å½•
        if "messages" not in st.session_state:
            st.session_state.messages = []

        # å±•ç¤ºå†å²å¯¹è¯
        for msg in st.session_state.messages:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        # æ¥æ”¶ç”¨æˆ·è¾“å…¥ (å¯¹è¯æ¡†åœ¨è¿™é‡Œï¼)
        if prompt := st.chat_input("é—®æˆ‘ï¼šè°çš„å¥–é‡‘æœ€é«˜ï¼Ÿ"):
            st.session_state.messages.append({"role": "user", "content": prompt})
            with st.chat_message("user"):
                st.markdown(prompt)

            with st.chat_message("assistant"):
                with st.spinner("AI æ­£åœ¨æ€è€ƒ..."):
                    # è°ƒç”¨ä¸Šé¢å®šä¹‰å¥½çš„å‡½æ•°
                    response = chat_with_ai(prompt, context_text)
                    st.markdown(response)
                    st.session_state.messages.append({"role": "assistant", "content": response})
    else:
        st.info("ğŸ’¡ è¯·å…ˆä¸Šä¼  Excel æ–‡ä»¶å¹¶å®Œæˆç»“ç®—ï¼ŒAI åŠ©æ‰‹å°†è‡ªåŠ¨å¼€å¯ã€‚")
# --- æ ¸å¿ƒ AI å‡½æ•°ï¼šç¡®ä¿å·¦ä¾§æ²¡æœ‰ä»»ä½•ç©ºæ ¼ï¼Œé¡¶æ ¼å†™ ---
def chat_with_ai(user_prompt, context_data):
    try:
        api_key = st.secrets["DEEPSEEK_API_KEY"]
        url = "https://api.deepseek.com/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }

        # 1. å…ˆå®šä¹‰äººè®¾æŒ‡ä»¤å˜é‡
        system_prompt = (
            "ä½ æ˜¯101ä¿±ä¹éƒ¨ä¸“å±çš„ã€é¦–å¸­è´¢åŠ¡å®¡è®¡å®˜ã€‘ã€‚ä½ çš„ç›®æ ‡æ˜¯åŸºäºæä¾›çš„ç»“ç®—æŠ¥è¡¨ï¼Œç»™å‡ºæå…·ä¸“ä¸šæ€§çš„è´¢åŠ¡æ´å¯Ÿã€‚\n\n"
            "## æ ¸å¿ƒæŠ€èƒ½ï¼š\n"
            "1. **è®¡ç®—ROIä¸æ•ˆèƒ½**ï¼šé€šè¿‡ï¼ˆé‡‘é¢ / æ’­æ”¾é‡ï¼‰è®¡ç®—æ¯ä¸‡æ¬¡æ’­æ”¾çš„æˆæœ¬ï¼Œè¯†åˆ«è°æ˜¯â€˜é«˜æ€§ä»·æ¯”åˆ›ä½œè€…â€™ã€‚\n"
            "2. **å¼‚å¸¸æ•°æ®è¯†åˆ«**ï¼šæŒ‡å‡ºæ’­æ”¾é‡æé«˜ä½†å¥–é‡‘æä½ï¼Œæˆ–é‡‘é¢ä¸æ’­æ”¾é‡æ¯”ä¾‹ä¸¥é‡å¤±è¡¡çš„æ¡ˆä¾‹ã€‚\n"
            "3. **è¶‹åŠ¿æ€»ç»“**ï¼šå¿«é€Ÿæ¦‚æ‹¬å“ªäº›å¹³å°çš„è¡¨ç°æ›´ç¬¦åˆå½“å‰çš„å¥–åŠ±æ”¿ç­–ã€‚\n\n"
            "## æ²Ÿé€šåŸåˆ™ï¼š\n"
            "- **æ•°æ®é©±åŠ¨**ï¼šä¸¥ç¦è§£é‡Šåè¯ï¼Œå¿…é¡»ç›´æ¥å¼•ç”¨æŠ¥è¡¨ä¸­çš„å…·ä½“æ•°å­—ã€‚\n"
            "- **ç®€æ´ä¸“ä¸š**ï¼šå¤šç”¨ç»“è®ºæ€§çŸ­è¯­ï¼Œå¦‚â€˜æ•°æ®å€’æŒ‚â€™ã€â€˜è½¬åŒ–æ•ˆç‡æœ€é«˜â€™ã€‚"
        )

        # 2. æ„é€ è¯·æ±‚è½½ä½“ (æ³¨æ„è¿™é‡Œçš„ payload åªå®šä¹‰ä¸€æ¬¡ï¼Œä¸”æ‹¬å·å®Œå…¨åŒ¹é…)
        payload = {
            "model": "deepseek-chat", 
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"ã€å½“å‰ç»“ç®—æŠ¥è¡¨æ•°æ®ã€‘ï¼š\n{context_data}\n\nã€ç”¨æˆ·æé—®ã€‘ï¼š{user_prompt}"}
            ],
            "temperature": 0.3
        }

        # 3. å‘é€è¯·æ±‚å¹¶è§£æç»“æœ
        response = requests.post(url, json=payload, headers=headers)
        res_json = response.json()
        
        if response.status_code != 200:
            return f"AI æš‚æ—¶æ‰çº¿äº† (APIé”™è¯¯): {res_json.get('error', {}).get('message', 'æœªçŸ¥é”™è¯¯')}"
            
        return res_json['choices'][0]['message']['content']

    except Exception as e:
        return f"AI æš‚æ—¶æ‰çº¿äº† (ç³»ç»Ÿé”™è¯¯): {str(e)}"
    render()
